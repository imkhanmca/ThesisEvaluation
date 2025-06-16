# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import os
import openai
import openpyxl
from docx import Document
from openai import OpenAI


openai_api_key = os.getenv('C:/Users/imran.khan/my-python-project2/myProject/venv/.env')
if openai_api_key:
    print(f"OpenAI API Key exists and begins {openai_api_key[:8]}")
else:
    print("OpenAI API Key not set - please head to the troubleshooting guide in the setup folder")

doc = Document('C:/Users/imran.khan/Desktop/Polytechnic/Feb - 2025/CLP/Thesis/202000739 Aesha Mohamed_10993646_assignsubmission_file/202000739.docx')
start_heading = "Abstract"
end_heading = "Acknowledgements"
capture = False
captured_paragraphs = []

workbook  = openpyxl.load_workbook('C:/Users/imran.khan/Downloads/OneDrive_1_6-1-2025/Aesha.xlsx')
worksheet = workbook.active
for row in worksheet.iter_rows():
    for cell in row:
        #print(cell.row, ' & ' , cell.column, ' = ', cell.value)
       if (cell.value=="Introductory contents (includes abstract)"):
           cell_value = worksheet.cell(row=cell.row, column=cell.column+1).value
           print(cell_value, end=' \n Word File Content:\n')
           for para in doc.paragraphs:
               text = para.text.strip()

               # Check if current paragraph is the start heading
               if text.lower() == start_heading.lower():
                   capture = True
                   continue  # Skip the heading itself

               # Stop capturing when end heading is reached
               if text.lower() == end_heading.lower():
                   capture = False
                   break  # Exit if you only want the first block between start and end

               # Capture paragraphs between headings
               if capture and text:
                   captured_paragraphs.append(text)

           # Print the result
           #print("Paragraphs between headings:")
           para = ""
           for p in captured_paragraphs:
               para = para + p
               #print("-" * 50)
           print(para)
           break
prompt = "Evaluate Abstract of an Academic project out of 5 Marks for the given Abstract:" + para + " Based on this Rubric:" + cell_value + " Grade this out of 5 Marks along with comments if any deductions first it should give only marks without any characters and than comments."
'''response = openai.ChatCompletion.create(
    model="gpt-3.5-turbo",  # or "gpt-4" if you have access
    messages=[
        {"role": "system", "content": "You are a helpful assistant."},
        {"role": "user", "content": prompt}
    ]
)

# Extract and print the response
reply = response['choices'][0]['message']['content']
print("ChatGPT response:")
print(reply)
'''
openai = OpenAI()
messages = [{"role": "user", "content": prompt}]
response = openai.chat.completions.create(
    model="gpt-4o-mini",
    messages=messages
)
#print(response.choices[0].message.content)
print(response.choices[0].message.content[0], '\n', response.choices[0].message.content[-1])